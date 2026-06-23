"""Build the multi-tab .xlsx audit, mirroring the existing audit format."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="0000FF")  # blue header, matches existing audits
WRAP = Alignment(wrap_text=True, vertical="top")
PRIORITY_FILL = {
    "Critical": PatternFill("solid", fgColor="F4CCCC"),
    "High": PatternFill("solid", fgColor="FCE5CD"),
    "Medium": PatternFill("solid", fgColor="FFF2CC"),
    "Low": PatternFill("solid", fgColor="D9EAD3"),
}


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


_IMAGE_KEYS = {"image_large"}


def build(path, client, rows, notes, df_raw, evidence_tabs,
          total_pages=None, total_images=None):
    wb = Workbook()

    # --- Observation tab ---
    ws = wb.active
    ws.title = "Observation"
    COUNT_HDR = "Count ⚠ DELETE BEFORE SHARING"
    ws.append(["Category", "Observation", "Priority", "Impact", "Reference", COUNT_HDR])
    _style_header(ws, 6)
    # Style count header in red so it's obvious
    ws.cell(row=1, column=6).font = Font(bold=True, color="FF0000")
    for r in rows:
        count = r.get("count")
        key = r.get("key", "")
        if count and key in _IMAGE_KEYS and total_images:
            count_label = f"{count} / {total_images} images"
        elif count and total_pages:
            count_label = f"{count} / {total_pages} pages"
        elif count:
            count_label = str(count)
        else:
            count_label = ""
        ws.append([r.get("category", ""), r["observation"], r["priority"],
                   r["impact"], r["reference"], count_label])
        row_i = ws.max_row
        fill = PRIORITY_FILL.get(r["priority"])
        if fill:
            ws.cell(row=row_i, column=3).fill = fill
        for c in range(1, 7):
            ws.cell(row=row_i, column=c).alignment = WRAP
    if notes:
        ws.append([])
        ws.append(["Notes:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True)
        for n in notes:
            ws.append([n])
            ws.cell(row=ws.max_row, column=1).alignment = WRAP
    _autosize(ws, [20, 66, 12, 55, 40, 22])
    ws.freeze_panes = "A2"

    # --- Evidence tabs ---
    for tab_name, headers, data_rows in evidence_tabs:
        wsx = wb.create_sheet(tab_name[:31])
        wsx.append(headers)
        _style_header(wsx, len(headers))
        for dr in data_rows:
            wsx.append(list(dr))
        _autosize(wsx, [min(90, max(18, len(str(h)) + 6)) for h in headers])
        wsx.freeze_panes = "A2"

    # --- Raw SF Internal all ---
    wsr = wb.create_sheet("SF Internal all")
    wsr.append(list(df_raw.columns))
    _style_header(wsr, len(df_raw.columns))
    for _, row in df_raw.iterrows():
        wsr.append(["" if pd.isna(v) else v for v in row.tolist()])
    wsr.freeze_panes = "A2"

    wb.save(path)
    return path
